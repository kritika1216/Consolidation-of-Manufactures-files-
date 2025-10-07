import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ✅ Only change this (keep folder in same location as this script)
FOLDER_PATH = "Consolidated Portfolios"
OUTPUT_FILE = "Consolidated_Output_SCM1.xlsx"
OUTPUT_SHEETNAME = "Consolidated"


def detect_header_row(file_path, sheet_name='Mapped Sheet'):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
    for idx, row in preview.iterrows():
        if any(str(cell).strip() == 'M.Item Name' for cell in row if pd.notna(cell)):
            return idx
    raise ValueError(f"'M.Item Name' not found in top 10 rows of: {file_path}")


def clean_columns(df):
    df.columns = df.columns.str.strip()
    for col in df.columns:
        if "volume share" in col.lower():
            df.rename(columns={col: "Volume Share %"}, inplace=True)
        elif "volume" in col.lower() and "share" not in col.lower():
            df.rename(columns={col: "Volume"}, inplace=True)
    return df


def load_and_prepare(file_path, manufacturer_name, sheet_name='Mapped Sheet'):
    header_row = detect_header_row(file_path, sheet_name)
    # No dtype argument, so both numbers and text are preserved in all columns
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    df = clean_columns(df)

    if 'M.Item Name' not in df.columns:
        raise ValueError(f"'M.Item Name' not found in {file_path}")

    df = df[df['M.Item Name'].notna() & (df['M.Item Name'].astype(str).str.strip() != '')]
    df['Manufacturer'] = manufacturer_name
    return df


def apply_header_colors(filepath, sheetname):
    wb = load_workbook(filepath)
    ws = wb[sheetname]
    light_orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    header_row = list(ws[1])
    m_item_indices = [idx for idx, cell in enumerate(header_row) if cell.value and str(cell.value).strip() == "M.Item Name"]
    m_idx = m_item_indices[0] if m_item_indices else None

    for idx, cell in enumerate(header_row):
        if cell.value and str(cell.value).strip() == "M.Item Name":
            cell.fill = light_green
        elif m_idx is not None:
            if idx < m_idx:
                cell.fill = light_orange
            else:
                cell.fill = light_green
        else:
            cell.fill = light_orange  # fallback
    wb.save(filepath)


def extract_manufacturer_name(file_name):
    return file_name.split('-')[0].strip()


def main():
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ Folder '{FOLDER_PATH}' not found. Please create the folder and add Excel files.")
        return

    all_xlsx = [f for f in os.listdir(FOLDER_PATH) if f.endswith('.xlsx')]
    if not all_xlsx:
        print(f"⚠️ No .xlsx files found in '{FOLDER_PATH}'")
        return

    all_data = []
    for file in all_xlsx:
        full_path = os.path.join(FOLDER_PATH, file)
        manufacturer = extract_manufacturer_name(file)
        try:
            df = load_and_prepare(full_path, manufacturer)
            all_data.append(df)
            print(f"✅ Processed: {file}")
        except Exception as e:
            print(f"❌ Skipped {file}, reason: {e}")

    if not all_data:
        print("⛔ No valid data to save.")
        return

    # Concatenate all DataFrames, Manufacturer is retained
    consolidated_df = pd.concat(all_data, ignore_index=True)

    # Drop the last column before saving
    consolidated_df = consolidated_df.iloc[:, :-1]

    # Save to a single sheet
    consolidated_df.to_excel(OUTPUT_FILE, sheet_name=OUTPUT_SHEETNAME, index=False)

    # Color header row
    apply_header_colors(OUTPUT_FILE, OUTPUT_SHEETNAME)

    print(f"\n✅ All done! Final Excel file saved as: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
