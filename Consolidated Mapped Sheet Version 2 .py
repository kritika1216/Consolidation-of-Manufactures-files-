import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math


FOLDER_PATH = "Final RFQ from MFG_MPUH"
OUTPUT_FILE = "consolidated_output2.xlsx"
OUTPUT_SHEETNAME = "Consolidated"


def detect_header_row(file_path, sheet_name='Mapped Sheet'):
    """Find header row by searching for 'M.Item Name'."""
    preview = None
    try:
        # Increase the number of rows to check for the header
        preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=50)

    except Exception as e:
        print(f"Error reading file {file_path} or sheet '{sheet_name}' during header detection: {e}")
        raise # Re-raise the exception after printing

    if preview is not None:
        for idx, row in preview.iterrows():
            if any(str(cell).strip() == 'M.Item Name' for cell in row if pd.notna(cell)):
                return idx

        # If loop finishes without finding header, raise error
        raise ValueError(f"'M.Item Name' not found in top 50 rows of sheet '{sheet_name}' in file: {file_path}")
    else:
         raise RuntimeError(f"Failed to load preview data for {file_path}, sheet '{sheet_name}'. Cannot proceed with header detection.")


def custom_round(x):
    """Rounds a number to the nearest integer, rounding up for >= 0.5."""
    if pd.isna(x):
        return x
    return int(math.floor(x + 0.5))


def clean_columns(df):
    """Rename & unify columns per requirements, consolidate volume data, and drop unnecessary columns."""
    df.columns = df.columns.str.strip()
    colmap = {}
    cols_to_drop = []

    # Specific renamings requested by the user
    colmap['Amanta (Volume Share %)'] = 'Volume Share'
    colmap['Amanta (Volume )'] = 'Volume'

    # Apply specific renamings first to standardize column names
    df.rename(columns=colmap, inplace=True)

    # Identify all potential source columns for Volume (case-insensitive, looking for "volume" anywhere)
    volume_cols = [col for col in df.columns if "volume" in col.lower() and "volume share" not in col.lower()]
    volume_share_cols = [col for col in df.columns if "volume share" in col.lower()]

    print(f"--- Debug: Identified volume columns (after initial rename): {volume_cols} ---") # Debug print
    print(f"--- Debug: Identified volume share columns (after initial rename): {volume_share_cols} ---") # Debug print


    # Ensure the target 'Volume' column exists and is initialized to None/NaN
    volume_col_target = 'Volume'
    if volume_col_target not in df.columns:
        df[volume_col_target] = pd.NA # Use pandas NA for missing numeric data
    else:
        # Attempt to convert the existing Volume column to numeric, coercing errors
        df[volume_col_target] = pd.to_numeric(df[volume_col_target], errors='coerce')


    # Consolidate data from all identified volume columns into the target 'Volume' column
    # Prioritize non-null values from the source columns to fill the target 'Volume' column
    for col in volume_cols:
        # Check if the column exists before attempting to access it and is not the target column itself
        if col in df.columns and col != volume_col_target:
            # Coerce data in the source column to numeric, replacing errors with NaN
            source_col_numeric = pd.to_numeric(df[col], errors='coerce')
            print(f"--- Debug: Consolidating data from column '{col}'. First 5 values (numeric): {source_col_numeric.head().tolist()} ---") # Debug print

            # Use fillna to merge data: fill NaNs in the target column with non-NaN values from the source
            df[volume_col_target] = df[volume_col_target].fillna(source_col_numeric)

            # Mark other volume columns for dropping, unless it's the target column itself
            if col != volume_col_target:
                 cols_to_drop.append(col)
        elif col in df.columns and col == volume_col_target:
             # If the target column itself is in the list, ensure its data is numeric
             df[volume_col_target] = pd.to_numeric(df[volume_col_target], errors='coerce')


    print(f"--- Debug: After consolidation, first 5 values in '{volume_col_target}': {df[volume_col_target].head().tolist()} ---") # Debug print


    # Handle Volume Share - consolidate and apply percentage/rounding
    volume_share_col_target = 'Volume Share'
    if volume_share_col_target not in df.columns:
        df[volume_share_col_target] = None

    # Consolidate all identified volume share columns into the target 'Volume Share' column
    for col in volume_share_cols:
         if col in df.columns and col != volume_share_col_target:
             source_col_numeric = pd.to_numeric(df[col], errors='coerce')
             df[volume_share_col_target] = df[volume_share_col_target].fillna(source_col_numeric)
             if col != volume_share_col_target:
                 cols_to_drop.append(col)
         elif col in df.columns and col == volume_share_col_target:
             df[volume_share_col_target] = pd.to_numeric(df[volume_share_col_target], errors='coerce')


    # Apply percentage conversion and rounding to the final 'Volume Share' column
    if volume_share_col_target in df.columns:
        # Multiply by 100 for percentage, coercing errors
        df[volume_share_col_target] = pd.to_numeric(df[volume_share_col_target], errors='coerce') * 100
        # Apply custom rounding
        df[volume_share_col_target] = df[volume_share_col_target].apply(custom_round)


    # Handle "Projected MFS Annual Qty" renaming
    if "Projected MFS Annual Qty" in df.columns:
         colmap["Projected MFS Annual Qty"] = "Projected MFS Annual Qty Unit Level"
         # Need to apply this renaming if not already done in the initial renaming step
         df.rename(columns={"Projected MFS Annual Qty": "Projected MFS Annual Qty Unit Level"}, inplace=True)


    # Apply any remaining renamings (should be handled by the initial rename now)
    # df.rename(columns=colmap, inplace=True) # This line is redundant now


    # Drop any columns that were marked for dropping, ensuring we don't drop the ones we just renamed
    # Also ensure we don't drop the target 'Volume' and 'Volume Share' columns themselves
    final_cols_to_drop = [c for c in cols_to_drop if c in df.columns and c not in [volume_col_target, volume_share_col_target]]
    df.drop(columns=final_cols_to_drop, errors='ignore', inplace=True)

    # Ensure required columns exist after cleaning
    for col in [volume_share_col_target, volume_col_target, "Projected MFS Annual Qty Unit Level", "M.Item Name"]:
        if col not in df.columns:
            df[col] = None # Add missing columns as None


    return df


def insert_after(df, col_name, new_cols):
    """Insert one or more new columns after col_name."""
    cols = list(df.columns)
    if col_name not in cols:
        return df
    idx = cols.index(col_name)
    # Ensure new_cols are not already in df.columns before inserting
    new_cols_to_insert = [nc for nc in new_cols if nc not in cols]
    if not new_cols_to_insert:
        return df # Nothing to insert

    # Create a new list of columns including the new ones
    new_cols_list = cols[:idx+1]
    new_cols_list.extend(new_cols_to_insert)
    new_cols_list.extend(cols[idx+1:])

    # Reindex the DataFrame with the new column order and add empty columns using .loc
    new_df = df.reindex(columns=new_cols_list)
    for nc in new_cols_to_insert:
        new_df.loc[:, nc] = ""  # empty column

    return new_df


def insert_custom_columns(df):
    # Insert "Form or Unit Type" after "Projected MFS Annual Qty Unit Level"
    if "Projected MFS Annual Qty Unit Level" in df.columns:
        df = insert_after(df, "Projected MFS Annual Qty Unit Level", ["Form or Unit Type"])
    # Insert "Potential at Pack Level" after "MFG Therapy Name"
    therapy_col = None
    for c in df.columns:
        if c.strip().lower() in ['mfg therapy name', 'therapy']:
            therapy_col = c
            break
    if therapy_col:
        df = insert_after(df, therapy_col, ["Potential at Pack Level"])
    return df


def remove_unnamed_19(df):
    if 'Unnamed: 19' in df.columns:
        df = df.drop('Unnamed: 19', axis=1)
    return df


def load_and_prepare(file_path, manufacturer_name, sheet_name='Mapped Sheet'):
    header_row = detect_header_row(file_path, sheet_name)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    df = clean_columns(df) # Clean and drop manufacturer-specific volume columns
    if 'M.Item Name' not in df.columns:
        # If M.Item Name is lost after cleaning, it means the cleaning logic was too aggressive
        # or the original header detection was wrong. Re-raise with more context.
        raise ValueError(f"'M.Item Name' not found in dataframe columns after cleaning in file: {file_path}. Original columns: {list(df.columns)}")

    df = insert_custom_columns(df)
    df = remove_unnamed_19(df)
    df = df[df['M.Item Name'].notna() & (df['M.Item Name'].astype(str).str.strip() != '')]
    df['Manufacturer'] = manufacturer_name # Add Manufacturer column
    return df


def apply_header_colors(filepath, sheetname, header_row_index):
    wb = load_workbook(filepath)
    ws = wb[sheetname]

    light_orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    highlight_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Columns to highlight in yellow (only inserted custom columns)
    newly_added = [
        'Form or Unit Type',
        'Potential at Pack Level',
    ]

    # Use the correct header row index (1-based for openpyxl)
    # Add 1 to the 0-based pandas index
    header_row_cells = list(ws[header_row_index + 1])

    # Find the index of the 'Volume' column
    volume_idx = None
    for idx, cell in enumerate(header_row_cells):
        if cell.value and str(cell.value).strip() == "Volume":
            volume_idx = idx
            break

    for idx, cell in enumerate(header_row_cells):
        value = "" if cell.value is None else str(cell.value).strip()
        # Highlight NEW columns in yellow
        if value in newly_added:
            cell.fill = highlight_yellow
        elif volume_idx is not None:
            # Compare the current column index with the index of 'Volume'
            if idx <= volume_idx: # Include Volume column in orange
                cell.fill = light_orange
            else:
                cell.fill = light_green
        else:
             # If Volume column is not found, apply orange to all
            cell.fill = light_orange

    wb.save(filepath)


def extract_manufacturer_name(file_name):
    return file_name.split('-')[0].strip()


def main():
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ Folder '{FOLDER_PATH}' not found. Please create the folder and add Excel files.")
        return

    all_xlsx = [f for f in os.listdir(FOLDER_PATH) if f.endswith('.xlsx')]
    if not all_xlsx:
        print(f"⚠️ No .xlsx files found in '{FOLDER_PATH}'")
        return

    all_data = []
    # Store header row index for coloring the output file
    output_header_row_index = 0 # Assuming header is always row 0 in the consolidated dataframe

    for file in all_xlsx:
        full_path = os.path.join(FOLDER_PATH, file)
        manufacturer = extract_manufacturer_name(file)
        try:
            print(f"--- Processing file: {file} ---")
            # Detect header row before loading and preparing the dataframe
            # Note: The header index detected here is 0-based for pandas read_excel
            # We don't need to store this for coloring the *output* file,
            # as the output header is always at index 0 of the consolidated dataframe.
            header_row_index_pandas = detect_header_row(full_path)
            df = load_and_prepare(full_path, manufacturer)
            all_data.append(df)
            print(f"✅ Successfully processed: {file}")
        except Exception as e:
            print(f"❌ Skipped {file}, reason: {type(e).__name__}: {e}") # Added more specific error info

    if not all_data:
        print("⛔ No valid data to save.")
        return

    # Concatenate dataframes
    consolidated_df = pd.concat(all_data, ignore_index=True)

    # Reorder columns to place 'Manufacturer' at the beginning and 'Volume' after 'Volume Share'
    current_cols = consolidated_df.columns.tolist()
    ordered_cols = ['Manufacturer']
    if 'Volume Share' in current_cols:
        ordered_cols.append('Volume Share')
        if 'Volume' in current_cols:
            ordered_cols.append('Volume')
    # Add other columns in their original order, excluding those already added
    ordered_cols.extend([col for col in current_cols if col not in ordered_cols])

    # Ensure all columns are in the correct order, handling potential missing columns
    final_ordered_cols = []
    for col in ordered_cols:
        if col in consolidated_df.columns:
            final_ordered_cols.append(col)
    # Add any remaining columns not in the ordered list (shouldn't happen with the current logic, but safe)
    final_ordered_cols.extend([col for col in consolidated_df.columns if col not in final_ordered_cols])


    consolidated_df = consolidated_df[final_ordered_cols]


    consolidated_df.to_excel(OUTPUT_FILE, sheet_name=OUTPUT_SHEETNAME, index=False)
    # Pass the header row index for coloring the output file (always 0 for the dataframe header)
    apply_header_colors(OUTPUT_FILE, OUTPUT_SHEETNAME, output_header_row_index)

    print(f"\n✅ All done! Final Excel file saved as: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()