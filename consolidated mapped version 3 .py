import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math


FOLDER_PATH = "final_rfq_from_mfg's_olive"
OUTPUT_FILE = "consolidated_Olive_file.xlsx"
OUTPUT_SHEETNAME = "Consolidated" 


def detect_header_row(file_path, sheet_name='Mapped Sheet'):
    """Find header row by searching for 'M.Item Name'."""
    preview = None
    try:
        # Increase the number of rows to check for the header
        preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=50)

    except Exception as e:
        print(f"Error reading file {file_path} or sheet '{sheet_name}' during header detection: {e}")
        raise # Re-raise the exception after printing

    if preview is not None:
        for idx, row in preview.iterrows():
            if any(str(cell).strip() == 'M.Item Name' for cell in row if pd.notna(cell)):
                return idx

        # If loop finishes without finding header, raise error
        raise ValueError(f"'M.Item Name' not found in top 50 rows of sheet '{sheet_name}' in file: {file_path}")
    else:
         raise RuntimeError(f"Failed to load preview data for {file_path}, sheet '{sheet_name}'. Cannot proceed with header detection.")


def custom_round(x):
    """Rounds a number to the nearest integer, rounding up for >= 0.5."""
    if pd.isna(x):
        return x
    return int(math.floor(x + 0.5))


def clean_columns(df):
    """Rename & unify columns per requirements, consolidate volume data, and drop unnecessary columns."""
    df.columns = df.columns.str.strip()
    colmap = {}
    cols_to_drop = []

    # Specific renamings requested by the user
    colmap['Amanta (Volume Share %)'] = 'Volume Share'
    colmap['Amanta (Volume )'] = 'Volume'
    # Add mappings for the columns the user wants to fix
    colmap['Projected MFS Annual Qty at Unit level'] = 'Projected MFS Annual Qty Unit Level'
    colmap['FORM OR UNIT TYPE'] = 'Form or Unit Type'
    colmap['MFG Therapy Name'] = 'MFG Therapy Name' # Explicitly keep this
    colmap['Potential at pack'] = 'Potential at Pack Level'


    # Apply specific renamings first to standardize column names
    df.rename(columns=colmap, inplace=True)

    # Identify all potential source columns for Volume (case-insensitive, looking for "volume" anywhere)
    volume_cols = [col for col in df.columns if "volume" in col.lower() and "volume share" not in col.lower()]
    volume_share_cols = [col for col in df.columns if "volume share" in col.lower()]

    print(f"--- Debug: Identified volume columns (after initial rename): {volume_cols} ---") # Debug print
    print(f"--- Debug: Identified volume share columns (after initial rename): {volume_share_cols} ---") # Debug print


    # Ensure the target 'Volume' column exists and is initialized to None/NaN
    volume_col_target = 'Volume'
    if volume_col_target not in df.columns:
        df[volume_col_target] = pd.NA # Use pandas NA for missing numeric data
    else:
        # Attempt to convert the existing Volume column to numeric, coercing errors
        df[volume_col_target] = pd.to_numeric(df[volume_col_target], errors='coerce')


    # Consolidate data from all identified volume columns into the target 'Volume' column
    # Prioritize non-null values from the source columns to fill the target 'Volume' column
    for col in volume_cols:
        # Check if the column exists before attempting to access it and is not the target column itself
        if col in df.columns and col != volume_col_target:
            # Coerce data in the source column to numeric, replacing errors with NaN
            source_col_numeric = pd.to_numeric(df[col], errors='coerce')
            print(f"--- Debug: Consolidating data from column '{col}'. First 5 values (numeric): {source_col_numeric.head().tolist()} ---") # Debug print

            # Use fillna to merge data: fill NaNs in the target column with non-NaN values from the source
            df[volume_col_target] = df[volume_col_target].fillna(source_col_numeric)

            # Mark other volume columns for dropping, unless it's the target column itself
            if col != volume_col_target:
                 cols_to_drop.append(col)
        elif col in df.columns and col == volume_col_target:
             # If the target column itself is in the list, ensure its data is numeric
             df[volume_col_target] = pd.to_numeric(df[volume_col_target], errors='coerce')


    print(f"--- Debug: After consolidation, first 5 values in '{volume_col_target}': {df[volume_col_target].head().tolist()} ---") # Debug print


    # Handle Volume Share - consolidate and apply percentage/rounding
    volume_share_col_target = 'Volume Share'
    if volume_share_col_target not in df.columns:
        df[volume_share_col_target] = None

    # Consolidate all identified volume share columns into the target 'Volume Share' column
    for col in volume_share_cols:
         if col in df.columns and col != volume_share_col_target:
             source_col_numeric = pd.to_numeric(df[col], errors='coerce')
             df[volume_share_col_target] = df[volume_share_col_target].fillna(source_col_numeric)
             if col != volume_share_col_target:
                 cols_to_drop.append(col)
         elif col in df.columns and col == volume_share_col_target:
             df[volume_share_col_target] = pd.to_numeric(df[volume_share_col_target], errors='coerce')


    # Apply percentage conversion and rounding to the final 'Volume Share' column
    if volume_share_col_target in df.columns:
        # Multiply by 100 for percentage, coercing errors
        df[volume_share_col_target] = pd.to_numeric(df[volume_share_col_target], errors='coerce') * 100
        # Apply custom rounding
        df[volume_share_col_target] = df[volume_share_col_target].apply(custom_round)


    # Handle "Projected MFS Annual Qty" renaming (this was already handled by the initial colmap)
    # if "Projected MFS Annual Qty" in df.columns:
    #      colmap["Projected MFS Annual Qty"] = "Projected MFS Annual Qty Unit Level"
    #      # Need to apply this renaming if not already done in the initial renaming step
    #      df.rename(columns={"Projected MFS Annual Qty": "Projected MFS Annual Qty Unit Level"}, inplace=True)


    # Apply any remaining renamings (should be handled by the initial rename now)
    # df.rename(columns=colmap, inplace=True) # This line is redundant now


    # Drop any columns that were marked for dropping, ensuring we don't drop the ones we just renamed
    # Also ensure we don't drop the target 'Volume' and 'Volume Share' columns themselves
    # Also ensure we don't drop the columns the user wants to keep and populate:
    # 'Projected MFS Annual Qty Unit Level', 'Form or Unit Type', 'MFG Therapy Name', 'Potential at Pack Level'
    cols_to_keep = [
        volume_col_target,
        volume_share_col_target,
        'Projected MFS Annual Qty Unit Level',
        'Form or Unit Type',
        'MFG Therapy Name',
        'Potential at Pack Level',
        "M.Item Name" # Always keep M.Item Name
        ]

    final_cols_to_drop = [c for c in cols_to_drop if c in df.columns and c not in cols_to_keep]
    df.drop(columns=final_cols_to_drop, errors='ignore', inplace=True)

    # Ensure required columns exist after cleaning (including the ones the user wants to fix)
    for col in cols_to_keep:
        if col not in df.columns:
            df[col] = None # Add missing columns as None


    return df


def insert_after(df, col_name, new_cols):
    """Insert one or more new columns after col_name."""
    cols = list(df.columns)
    if col_name not in cols:
        return df
    idx = cols.index(col_name)
    # Ensure new_cols are not already in df.columns before inserting
    new_cols_to_insert = [nc for nc in new_cols if nc not in cols]
    if not new_cols_to_insert:
        return df # Nothing to insert

    # Create a new list of columns including the new ones
    new_cols_list = cols[:idx+1]
    new_cols_list.extend(new_cols_to_insert)
    new_cols_list.extend(cols[idx+1:])

    # Reindex the DataFrame with the new column order and add empty columns using .loc
    new_df = df.reindex(columns=new_cols_list)
    # Do not add empty columns here, as clean_columns should ensure they exist with data if found
    # The reindex itself will add them with NaN if they weren't in the original df columns

    return new_df


def insert_custom_columns(df):
    # Insert "Form or Unit Type" after "Projected MFS Annual Qty Unit Level"
    # These columns should now be handled by clean_columns and reindexing in main
    # This function might become redundant if the desired order is handled by reindex in main
    # Let's keep it for now but ensure it doesn't overwrite data.
    # The reindex in main will set the final order and add missing columns as NaN.
    return df # Returning df without modification for now


def remove_unnamed_19(df):
    if 'Unnamed: 19' in df.columns:
        df = df.drop('Unnamed: 19', axis=1)
    return df


def load_and_prepare(file_path, manufacturer_name, sheet_name='Mapped Sheet'):
    header_row = detect_header_row(file_path, sheet_name)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    df = clean_columns(df) # Clean and drop manufacturer-specific volume columns
    if 'M.Item Name' not in df.columns:
        # If M.Item Name is lost after cleaning, it means the cleaning logic was too aggressive
        # or the original header detection was wrong. Re-raise with more context.
        raise ValueError(f"'M.Item Name' not found in dataframe columns after cleaning in file: {file_path}. Original columns: {list(df.columns)}")

    # insert_custom_columns is now handled by reindexing in main, no need to call here
    # df = insert_custom_columns(df)
    df = remove_unnamed_19(df)
    df = df[df['M.Item Name'].notna() & (df['M.Item Name'].astype(str).str.strip() != '')]
    df['Manufacturer'] = manufacturer_name # Add Manufacturer column
    return df


def apply_header_colors(filepath, sheetname, header_row_index):
    wb = load_workbook(filepath)
    ws = wb[sheetname]

    light_orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    highlight_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Columns to highlight in yellow (only inserted custom columns)
    # These should now be the ones the user explicitly mentioned are missing
    newly_added_or_fixed = [
        'Projected MFS Annual Qty Unit Level',
        'Form or Unit Type',
        'Potential at Pack Level',
        'MFG Therapy Name' # Also highlight this as it was mentioned
    ]

    # Use the correct header row index (1-based for openpyxl)
    # Add 1 to the 0-based pandas index
    header_row_cells = list(ws[header_row_index + 1])

    # Find the index of the 'Volume' column
    volume_idx = None
    for idx, cell in enumerate(header_row_cells):
        if cell.value and str(cell.value).strip() == "Volume":
            volume_idx = idx
            break

    for idx, cell in enumerate(header_row_cells):
        value = "" if cell.value is None else str(cell.value).strip()
        # Highlight NEW/FIXED columns in yellow
        if value in newly_added_or_fixed:
            cell.fill = highlight_yellow
        elif volume_idx is not None:
            # Compare the current column index with the index of 'Volume'
            if idx <= volume_idx: # Include Volume column in orange
                cell.fill = light_orange
            else:
                cell.fill = light_green
        else:
             # If Volume column is not found, apply orange to all
            cell.fill = light_orange

    wb.save(filepath)


def extract_manufacturer_name_from_index(file_path):
    """Extracts manufacturer name from 'Index' sheet, cell A9 using openpyxl."""
    print(f"--- Attempting to extract manufacturer name from '{file_path}' 'Index' sheet, cell A9 ---")
    manufacturer_name = "Unknown Manufacturer"
    try:
        wb = load_workbook(file_path, read_only=True)
        if 'Index' in wb.sheetnames:
            ws = wb['Index']
            # Read cell A9 (row 9, column 1)
            cell_value = ws.cell(row=9, column=1).value
            print(f"--- Value found in cell A9: {cell_value} ---")
            if pd.notna(cell_value):
                manufacturer_name = str(cell_value).strip()
        else:
            print(f"⚠️ 'Index' sheet not found in file: {file_path}")

    except FileNotFoundError:
        print(f"❌ Error: File not found at {file_path}")
    except Exception as e:
        print(f"❌ Error extracting manufacturer name from {file_path} 'Index' sheet, cell A9: {e}")
        print(f"--- Specific Error details: {type(e).__name__}: {e} ---")

    print(f"--- Final extracted manufacturer name: {manufacturer_name} ---")
    return manufacturer_name


def main():
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ Folder '{FOLDER_PATH}' not found. Please create the folder and add Excel files.")
        return

    all_xlsx = [f for f in os.listdir(FOLDER_PATH) if f.endswith('.xlsx')]
    if not all_xlsx:
        print(f"⚠️ No .xlsx files found in '{FOLDER_PATH}'")
        return

    all_data = []
    # Store header row index for coloring the output file
    output_header_row_index = 0 # Assuming header is always row 0 in the consolidated dataframe

    for file in all_xlsx:
        full_path = os.path.join(FOLDER_PATH, file)
        # Extract manufacturer name from the 'INDEX' sheet
        manufacturer = extract_manufacturer_name_from_index(full_path)

        try:
            print(f"--- Processing file: {file} ---")
            # Detect header row before loading and preparing the dataframe
            # Note: The header index detected here is 0-based for pandas read_excel
            # We don't need to store this for coloring the *output* file,
            # as the output header is always at index 0 of the consolidated dataframe.
            header_row_index_pandas = detect_header_row(full_path)
            # Pass the extracted manufacturer name to load_and_prepare
            df = load_and_prepare(full_path, manufacturer)
            all_data.append(df)
            print(f"✅ Successfully processed: {file}")
        except Exception as e:
            print(f"❌ Skipped {file}, reason: {type(e).__name__}: {e}") # Added more specific error info

    if not all_data:
        print("⛔ No valid data to save.")
        return

    # Concatenate dataframes
    consolidated_df = pd.concat(all_data, ignore_index=True)

    # Define the desired final column order based on user's implicit order and requirements
    desired_columns = [
        "Manufacturer",
        "Hospital Name",
        "MFS",
        "Therapy",
        "Projected MFS Annual Qty Unit Level", # This should now be populated
        "Form or Unit Type", # This should now be populated
        "Volume Share",
        "Volume",
        "M.Item Name",
        "MFG Therapy Name", # This should now be populated
        "Potential at Pack Level", # This should now be populated
        "FORM OR UNIT TYPE BY AP",
        "UPP",
        "UPP BY AP",
        "MRP / Pack level",
        "Cost / Pack level",
        "MRP / Unit level",
        "Cost / Unit level",
        "GST%",
        "Quote Validity till date",
        "Scheme",
        "Scheme Validity till date",
        "Turn Over Discount",
        "TOD Validity till date",
    ]

    # Reindex the DataFrame to match the desired column order, adding missing columns as None
    # This will also handle the placement of the newly populated columns
    consolidated_df = consolidated_df.reindex(columns=desired_columns)

    # Ensure 'Therapy' column is populated if 'MFG Therapy Name' exists and 'Therapy' is None
    # This logic seems correct and should remain
    if 'Therapy' in consolidated_df.columns and 'MFG Therapy Name' in consolidated_df.columns:
        consolidated_df['Therapy'] = consolidated_df['Therapy'].fillna(consolidated_df['MFG Therapy Name'])

    # Ensure Volume Share is numeric and rounded after reindexing
    # This logic seems correct and should remain
    volume_share_col_target = 'Volume Share'
    if volume_share_col_target in consolidated_df.columns:
         consolidated_df[volume_share_col_target] = pd.to_numeric(consolidated_df[volume_share_col_target], errors='coerce')
         # Apply percentage conversion and custom rounding only if the data is numeric
         if pd.api.types.is_numeric_dtype(consolidated_df[volume_share_col_target]):
            consolidated_df[volume_share_col_target] = consolidated_df[volume_share_col_target] # Remove the division by 100
            # The rounding should ideally happen in clean_columns. Let's trust that for now.


    consolidated_df.to_excel(OUTPUT_FILE, sheet_name=OUTPUT_SHEETNAME, index=False)
    # Pass the header row index for coloring the output file (always 0 for the dataframe header)
    apply_header_colors(OUTPUT_FILE, OUTPUT_SHEETNAME, output_header_row_index)

    print(f"\n✅ All done! Final Excel file saved as: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()